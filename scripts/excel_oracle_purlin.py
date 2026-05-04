#!/usr/bin/env python3
"""
Oracle for the purlin (прогоны ЛСТК) calculator.

Loads the source Excel (v2.0), patches input cells with each test scenario,
recalculates via `libreoffice --headless --calc`, then reads the calculated
purlin selection cells (Лист1!D63:K70) for both МП350 and МП390 grades.

Output: JSON to stdout with results per scenario.
"""

from __future__ import annotations

import json
import os
import shutil
import subprocess
import sys
import tempfile
from typing import Any

import openpyxl

SRC_XLSX = os.environ.get(
    "PURLIN_XLSX",
    os.path.expanduser("~/attachments/4499ab5f-a339-427e-bc8c-823833e3177d/+++v2.0.xlsx"),
)

# Scenarios ----------------------------------------------------------------

# Default (Уфа, snow=2.4, wind=0.6, span=24, h=12, length=60)
S_DEFAULT = {
    "name": "S1 default (Уфа, span=24, h=12)",
    "gamma_n": 1,
    "roof_shape": "двускатная",
    "span": 24,
    "length": 60,
    "height": 12,
    "slope": 6,
    "frame_pitch": 6,
    "fachwerk_pitch": 6,
    "terrain": "В",
    "w0": 0.6,
    "Sg": 2.4,
    "roof_struct": "С-П 150 мм",
    "snow_drift": "нет",
    "drift_drop_h": 4.5,
    "drift_existing": 9.5,
    "max_step_mm": 1500,
    "min_step_mm": 1500,
    "max_util": "по умолчанию",
    "snowguard": "нет",
    "fence": "нет",
}

S_2 = dict(S_DEFAULT, name="S2 span=18", span=18)
S_3 = dict(S_DEFAULT, name="S3 span=30", span=30)
S_4 = dict(S_DEFAULT, name="S4 наше 250 мм", roof_struct="наше 250 мм")
S_5 = dict(S_DEFAULT, name="S5 small snow Sg=1.2", Sg=1.2)
S_6 = dict(S_DEFAULT, name="S6 large wind w0=1.2", w0=1.2)
S_7 = dict(S_DEFAULT, name="S7 wide range 500..1500", min_step_mm=500, max_step_mm=1500)
S_8 = dict(S_DEFAULT, name="S8 K=0.7", max_util=70)

SCENARIOS = [S_DEFAULT, S_2, S_3, S_4, S_5, S_6, S_7, S_8]


def patch_and_recalc(out_dir: str, params: dict[str, Any], idx: int) -> dict[str, Any]:
    wb = openpyxl.load_workbook(SRC_XLSX, data_only=False)
    ws = wb["Лист1"]

    # Inputs
    ws["B3"] = params["gamma_n"]
    ws["B4"] = params["roof_shape"]
    ws["B7"] = params["span"]
    ws["B9"] = params["length"]
    ws["B10"] = params["height"]
    ws["B11"] = params["slope"]
    ws["B12"] = params["frame_pitch"]
    ws["B13"] = params["fachwerk_pitch"]
    ws["B16"] = params["terrain"]
    # Override formula-driven w0/Sg with literal numbers (bypass external lookups)
    ws["B17"] = params["w0"]
    ws["B18"] = params["Sg"]
    ws["B19"] = params["roof_struct"]
    ws["B22"] = params["snow_drift"]
    ws["B23"] = params["drift_drop_h"]
    ws["B24"] = params["drift_existing"]
    ws["B38"] = params["max_step_mm"]
    ws["B39"] = params["min_step_mm"]
    ws["B41"] = params.get("ties", "нет")
    ws["B44"] = params["snowguard"]
    ws["B45"] = params["fence"]
    ws["B63"] = params["max_util"]

    src = os.path.join(out_dir, f"in_{idx}.xlsx")
    wb.save(src)

    # Recalc with LibreOffice
    subprocess.run(
        ["libreoffice", "--headless", "--calc", "--convert-to", "xlsx", "--outdir", out_dir, src],
        check=True,
        capture_output=True,
    )
    out_path = os.path.join(out_dir, f"in_{idx}.xlsx")  # libreoffice overwrites the same name

    wb2 = openpyxl.load_workbook(out_path, data_only=True)
    ws2 = wb2["Лист1"]

    # Read purlin selection: МП350 in rows 63,64,65 (2ТПС, 2ПС, Z), МП390 in rows 68,69,70
    def read_sel(row: int) -> dict[str, Any]:
        """Reads D=profile_name, E=spacing_mm, F=mass_per_m, G=mass_per_frame_step, H=mass_building."""
        return {
            "profile": ws2.cell(row, 4).value,    # D
            "spacing": ws2.cell(row, 5).value,    # E
            "mass_per_m": ws2.cell(row, 6).value, # F
            "mass_per_frame": ws2.cell(row, 7).value, # G
            "mass_building": ws2.cell(row, 8).value,  # H
        }

    return {
        "name": params["name"],
        "params": params,
        "MP350_2TPS": read_sel(63),
        "MP350_2PS": read_sel(64),
        "MP350_Z": read_sel(65),
        "MP390_2TPS": read_sel(68),
        "MP390_2PS": read_sel(69),
        "MP390_Z": read_sel(70),
        # Also some loads for sanity check
        "loads": {
            "q_snow": ws2["D17"].value if isinstance(ws2["D17"].value, (int, float)) else None,
            "B25_mu2": ws2["B25"].value,
            "designSpan_B8": ws2["B8"].value,
        },
    }


def main() -> None:
    if not os.path.exists(SRC_XLSX):
        print(f"Source xlsx not found: {SRC_XLSX}", file=sys.stderr)
        sys.exit(1)

    out_dir = tempfile.mkdtemp(prefix="purlin_oracle_")
    print(f"# tmp dir: {out_dir}", file=sys.stderr)

    results = []
    for i, sc in enumerate(SCENARIOS):
        print(f"# {sc['name']}", file=sys.stderr)
        try:
            r = patch_and_recalc(out_dir, sc, i)
            results.append(r)
        except subprocess.CalledProcessError as e:
            print(f"# FAILED: {e.stderr.decode() if e.stderr else e}", file=sys.stderr)
            results.append({"name": sc["name"], "error": str(e)})

    print(json.dumps(results, ensure_ascii=False, indent=2, default=str))

    # Cleanup (disabled while debugging)
    # shutil.rmtree(out_dir, ignore_errors=True)


if __name__ == "__main__":
    main()
