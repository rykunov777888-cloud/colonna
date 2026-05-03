#!/usr/bin/env python3
"""
Use LibreOffice to recalculate the source Excel for each test
scenario and capture the actual N/M values it produces. This is
the ground-truth oracle for verifying our TypeScript engine.

Strategy:
  1. Open the source Excel via openpyxl, override the relevant
     input cells (Сводка!B7 etc.) with hard-coded values, save
     a temp copy.
  2. Convert that copy via `libreoffice --convert-to xlsx --calc`
     so the formulas get recalculated and cached.
  3. Re-read with openpyxl (data_only=True) and pull out
     B51 (N) and B54 (M).

We must bypass external lookups for w0/sg because the original
sheet pulls them via `[1]СП 20` from an external file. We set
B19/B20 to fixed values to force literal numeric inputs.
"""
import openpyxl
import subprocess
import shutil
import os
import sys
import tempfile
import json

SRC_XLSX = "/home/ubuntu/attachments/31cc5e15-2e68-4a08-998c-834f1b5a9fc6/+++v6.1+1.xlsx"


def patch_and_recalc(out_dir, params, idx):
    """Patch input cells, save to out_dir, recalculate via libreoffice."""
    wb = openpyxl.load_workbook(SRC_XLSX, data_only=False)
    ws = wb["Сводка"]

    # Bypass external lookups by hard-coding values:
    ws["B19"].value = params["w0"]    # ветровая w0
    ws["B20"].value = params["sg"]    # снеговая sg
    # also override the IF chain (just rewrite the cell value)

    # Geometry / loads
    ws["B7"] = params["span"]
    ws["B8"] = params["length"]
    ws["B9"] = params["h"]
    ws["B10"] = params["slope"]
    ws["B11"] = params["frame_pitch"]
    ws["B12"] = params["fachverk_pitch"]
    ws["B13"] = params["spans"]      # "один" / "более одного"
    ws["B14"] = "есть" if params["ties"] else "нет"
    ws["B18"] = params["terrain"]
    ws["B48"] = params["col_type_ru"]
    ws["B50"] = params["addition"]
    ws["B3"] = params["gamma_n"]

    # Force recalc on open
    wb.properties.calcMode = "auto"

    in_dir = os.path.join(out_dir, f"in_{idx}")
    out_sub = os.path.join(out_dir, f"out_{idx}")
    os.makedirs(in_dir, exist_ok=True)
    os.makedirs(out_sub, exist_ok=True)
    src_copy = os.path.join(in_dir, f"sc_{idx}.xlsx")
    wb.save(src_copy)

    # Recalculate via LibreOffice (headless). Output dir MUST differ from
    # input dir, otherwise LibreOffice errors with "verify input parameters".
    result = subprocess.run(
        [
            "libreoffice", "--headless", "--calc",
            "--convert-to", "xlsx",
            "--outdir", out_sub,
            src_copy,
        ],
        capture_output=True, text=True, timeout=120,
    )
    if result.returncode != 0:
        raise RuntimeError(f"libreoffice failed: {result.stderr}")

    out_file = os.path.join(out_sub, f"sc_{idx}.xlsx")
    wb2 = openpyxl.load_workbook(out_file, data_only=True)
    ws2 = wb2["Сводка"]
    return {
        "N": ws2["B51"].value,
        "M": ws2["B54"].value,
        "wind_h": ws2["B25"].value,
        "wind_v": ws2["C25"].value,
        "snow": ws2["B24"].value,
    }


SCENARIOS = [
    ("Excel default (фахв., w0=0.6, sg=1.7)",
     dict(span=40, length=80, h=11.5, slope=6, frame_pitch=6, fachverk_pitch=6,
          spans="один", ties=False, terrain="В", col_type_ru="фахверковая",
          w0=0.6, sg=1.7, gamma_n=1.0, addition=15)),
    ("Крайняя колонна, один пролёт, без связей",
     dict(span=40, length=80, h=11.5, slope=6, frame_pitch=6, fachverk_pitch=6,
          spans="один", ties=False, terrain="В", col_type_ru="крайняя",
          w0=0.6, sg=1.7, gamma_n=1.0, addition=15)),
    ("Средняя, связи=есть, более одного пролёта",
     dict(span=40, length=80, h=11.5, slope=6, frame_pitch=6, fachverk_pitch=6,
          spans="более одного", ties=True, terrain="В", col_type_ru="средняя",
          w0=0.6, sg=1.7, gamma_n=1.0, addition=15)),
    ("h=20м, фахверковая",
     dict(span=40, length=80, h=20, slope=6, frame_pitch=6, fachverk_pitch=6,
          spans="один", ties=False, terrain="В", col_type_ru="фахверковая",
          w0=0.6, sg=1.7, gamma_n=1.0, addition=15)),
    ("w0=0.85 (VII р.), фахверковая",
     dict(span=40, length=80, h=11.5, slope=6, frame_pitch=6, fachverk_pitch=6,
          spans="один", ties=False, terrain="В", col_type_ru="фахверковая",
          w0=0.85, sg=1.7, gamma_n=1.0, addition=15)),
    ("Sg=3.0 (VI р.), фахверковая",
     dict(span=40, length=80, h=11.5, slope=6, frame_pitch=6, fachverk_pitch=6,
          spans="один", ties=False, terrain="В", col_type_ru="фахверковая",
          w0=0.6, sg=3.0, gamma_n=1.0, addition=15)),
    ("Местность A, фахверковая",
     dict(span=40, length=80, h=11.5, slope=6, frame_pitch=6, fachverk_pitch=6,
          spans="один", ties=False, terrain="А", col_type_ru="фахверковая",
          w0=0.6, sg=1.7, gamma_n=1.0, addition=15)),
    ("Пролёт 60м, фахверковая",
     dict(span=60, length=80, h=11.5, slope=6, frame_pitch=6, fachverk_pitch=6,
          spans="один", ties=False, terrain="В", col_type_ru="фахверковая",
          w0=0.6, sg=1.7, gamma_n=1.0, addition=15)),
    ("Уклон 20°, фахверковая",
     dict(span=40, length=80, h=11.5, slope=20, frame_pitch=6, fachverk_pitch=6,
          spans="один", ties=False, terrain="В", col_type_ru="фахверковая",
          w0=0.6, sg=1.7, gamma_n=1.0, addition=15)),
    ("γₙ=1.1, крайн., связи, multi",
     dict(span=40, length=80, h=11.5, slope=6, frame_pitch=6, fachverk_pitch=6,
          spans="более одного", ties=True, terrain="В", col_type_ru="крайняя",
          w0=0.6, sg=1.7, gamma_n=1.1, addition=15)),
]


def main():
    if len(sys.argv) < 2:
        out_path = "/tmp/excel_oracle_results.json"
    else:
        out_path = sys.argv[1]

    results = []
    with tempfile.TemporaryDirectory() as tmp:
        for i, (label, params) in enumerate(SCENARIOS, 1):
            print(f"[{i}/{len(SCENARIOS)}] {label}")
            try:
                r = patch_and_recalc(tmp, params, i)
                print(f"   N={r['N']:.2f}  M={r['M']:.2f}  "
                      f"wind_h={r['wind_h']:.4f}  snow={r['snow']:.4f}")
                results.append({"label": label, "params": params, "result": r})
            except Exception as e:
                print(f"   ERROR: {e}")
                results.append({"label": label, "params": params, "error": str(e)})

    with open(out_path, "w") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)
    print(f"\nSaved → {out_path}")


if __name__ == "__main__":
    main()
